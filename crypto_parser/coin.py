from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import requests
import os.path
import time


coin_name = input('Введи название криптовалюты: ').lower().strip()


HEADERS = {
    'User-Agent': "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36"
}

URL = f'https://coinmarketcap.com/currencies/{coin_name}/'
coin = URL.split('/')[-2]
now = datetime.now()


def get_coin_course(URL):
    get_html = requests.get(URL, headers=HEADERS)
    soup = BeautifulSoup(get_html.content, 'html.parser')

    name_list = []
    price_list = []
    filter_list = []

    tables = soup.find('table')

    while tables is None:
        print('Вставлена неверная валюта, завершаю программу')
        exit()
    else:
        for table in tables.find_all('tbody'):
            rows = table.find_all('tr')
            for row in rows:
                text = row.find('th').text
                numbers = row.find_all('td')
                get_price_and_maxmin24 = row.find('td').text
                # закидываем удачные значения в фильтр, нынешную стоимость[0] и "макс / мин24"[2]
                filter_list.append(get_price_and_maxmin24)
                name_list.append(text)
                for number in numbers:
                    try:
                        percent = number.find('div').text
                        price = number.find('span').text
                        price_list.append(price)
                        price_list.append(percent)
                    except Exception as e:
                        pass

        if os.path.isfile(f'data_{coin}.xlsx'):
            wb = load_workbook(f'data_{coin}.xlsx')
            ws = wb.active
            print('Файл наден!')
            print(f'Открываю файл excel data_{coin}.xlsx')
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'data'
            print(f'Создаю файл под названием data_{coin}.xlsx')

        ws['A1'] = name_list[0]  # coin price
        ws['B1'] = name_list[1]  # coin change 24 hours
        ws['C1'] = name_list[2]  # coin min / max 24 hours
        ws['D1'] = 'Time'

        ws.append([filter_list[0].replace(".", ","), price_list[0].replace(
            ".", ","), filter_list[2].replace(".", ","), now.strftime('%H:%M-%d/%m/%Y')])

        print('Записываю данные')
        print(f'Сохраняю файл data_{coin}.xlsx в корневую папку')

        wb.save(f'data_{coin}.xlsx')
        print('Файл сохранен!')


while True:
    get_coin_course(URL)
    time_wait = 10 * 180  # * 360
    print(f'Жду {time_wait // 60} минут до следующего парсера.')   # // 3600 час
    time.sleep(time_wait)
